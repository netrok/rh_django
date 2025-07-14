from datetime import date, datetime

from django.db.models import Count
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils.timezone import now
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from rest_framework import generics, filters, permissions
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.exceptions import PermissionDenied, ValidationError
from weasyprint import HTML

from .models import Empleado, Bitacora, BitacoraEmpleado
from .permissions import IsRRHHOrAdmin, IsGerenteOrAdmin, IsSuperAdmin
from .serializers import (
    EmpleadoSerializer,
    BitacoraSerializer,
    BitacoraEmpleadoSerializer,
)
from .utils import (
    registrar_bitacora,
    registrar_exportacion_empleados,
    registrar_intento_fallido_exportacion,
)


# 📄 Listar y Crear empleados
class EmpleadoListCreateAPIView(generics.ListCreateAPIView):
    queryset = Empleado.objects.all().order_by('id')
    serializer_class = EmpleadoSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    permission_classes = [IsAuthenticated]

    filterset_fields = ['activo', 'puesto', 'departamento', 'genero', 'estado_civil', 'fecha_ingreso']
    search_fields = ['nombres', 'apellido_paterno', 'apellido_materno', 'num_empleado', 'curp', 'rfc', 'nss', 'email', 'telefono']
    ordering_fields = ['num_empleado', 'apellido_paterno', 'apellido_materno', 'fecha_ingreso', 'departamento', 'puesto']

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsAuthenticated(), IsRRHHOrAdmin()]
        return super().get_permissions()

    def perform_create(self, serializer):
        empleado = serializer.save()
        empleado._request_user = self.request.user
        registrar_bitacora(
            instancia=empleado,
            accion='CREACIÓN',
            usuario=self.request.user,
            cambios={'detalle': 'Empleado creado desde la API'}
        )


# 🔍 Consultar, Actualizar, Eliminar empleados
class EmpleadoRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Empleado.objects.all()
    serializer_class = EmpleadoSerializer

    def get_permissions(self):
        if self.request.method == 'DELETE':
            return [IsAuthenticated(), IsSuperAdmin()]
        elif self.request.method in ['PUT', 'PATCH']:
            return [IsAuthenticated(), IsGerenteOrAdmin()]
        return [IsAuthenticated()]

    def perform_update(self, serializer):
        instance = self.get_object()
        anterior_activo = instance.activo
        empleado = serializer.save()
        empleado._request_user = self.request.user

        if 'activo' in serializer.validated_data and serializer.validated_data['activo'] != anterior_activo:
            accion = 'ACTIVACIÓN' if serializer.validated_data['activo'] else 'INACTIVACIÓN'
            registrar_bitacora(
                instancia=empleado,
                accion=accion,
                usuario=self.request.user,
                cambios={'activo': f'{anterior_activo} → {empleado.activo}'}
            )
        else:
            registrar_bitacora(
                instancia=empleado,
                accion='EDICIÓN',
                usuario=self.request.user,
                instancia_anterior=instance
            )

    def perform_destroy(self, instance):
        if instance.bitacoraempleado_set.exists():
            raise ValidationError("❌ No se puede eliminar: este empleado tiene historial ligado en la bitácora.")

        instance._request_user = self.request.user
        registrar_bitacora(
            instancia=instance,
            accion='ELIMINACIÓN',
            usuario=self.request.user,
            cambios={'detalle': 'Empleado eliminado desde la API'}
        )
        instance.delete()


# 📊 Dashboard de estadísticas
class EmpleadoDashboardAPIView(APIView):
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    def get(self, request):
        hoy = date.today()
        empleados = Empleado.objects.all()
        total = empleados.count()
        activos = empleados.filter(activo=True).count()
        inactivos = empleados.filter(activo=False).count()

        empleados_con_edad = empleados.exclude(fecha_nacimiento=None)
        edades = [
            hoy.year - e.fecha_nacimiento.year -
            ((hoy.month, hoy.day) < (e.fecha_nacimiento.month, e.fecha_nacimiento.day))
            for e in empleados_con_edad
        ]
        edad_promedio = round(sum(edades) / len(edades), 1) if edades else 0

        return Response({
            "total_empleados": total,
            "activos": activos,
            "inactivos": inactivos,
            "edad_promedio": edad_promedio,
            "por_departamento": list(empleados.values('departamento').annotate(total=Count('id'))),
            "por_puesto": list(empleados.values('puesto').annotate(total=Count('id'))),
            "por_genero": list(empleados.values('genero').annotate(total=Count('id'))),
        })


# 🗕 Exportar a Excel
class EmpleadoExportExcelAPIView(APIView):
    permission_classes = [IsAuthenticated, IsRRHHOrAdmin]

    def get(self, request):
        registrar_exportacion_empleados(request, tipo_exportacion='Excel')

        wb = Workbook()
        ws = wb.active
        ws.title = "Empleados"

        columnas = [
            'ID', 'Número de empleado', 'Nombres', 'Apellido paterno', 'Apellido materno',
            'Fecha de nacimiento', 'Género', 'Estado civil', 'CURP', 'RFC', 'NSS', 'Teléfono',
            'Email', 'Puesto', 'Departamento', 'Fecha de ingreso', 'Activo',
        ]
        ws.append(columnas)

        for emp in Empleado.objects.all():
            ws.append([
                emp.id, emp.num_empleado, emp.nombres, emp.apellido_paterno, emp.apellido_materno,
                emp.fecha_nacimiento.strftime('%Y-%m-%d') if emp.fecha_nacimiento else "",
                emp.genero, emp.estado_civil, emp.curp, emp.rfc, emp.nss, emp.telefono,
                emp.email, emp.puesto, emp.departamento,
                emp.fecha_ingreso.strftime('%Y-%m-%d') if emp.fecha_ingreso else "",
                "Sí" if emp.activo else "No"
            ])

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        nombre_archivo = f"empleados_{now().date()}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        wb.save(response)
        return response

    def permission_denied(self, request, message=None, code=None):
        registrar_intento_fallido_exportacion(request, tipo_exportacion='Excel')
        raise PermissionDenied("No tienes permisos para exportar empleados.")


# 🧾 Exportar a PDF
class EmpleadoExportPdfAPIView(APIView):
    permission_classes = [IsAuthenticated, IsRRHHOrAdmin]

    def get(self, request):
        registrar_exportacion_empleados(request, tipo_exportacion='PDF')

        empleados = Empleado.objects.all()
        fecha = datetime.now().strftime('%d/%m/%Y %H:%M')

        html_string = render_to_string("empleados/empleados_pdf.html", {
            "empleados": empleados,
            "fecha": fecha,
            "empresa": "Mi Empresa S.A. de C.V.",
        })

        pdf = HTML(string=html_string).write_pdf()
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="empleados.pdf"'
        return response

    def permission_denied(self, request, message=None, code=None):
        registrar_intento_fallido_exportacion(request, tipo_exportacion='PDF')
        raise PermissionDenied("No tienes permisos para exportar empleados.")


# 📋 Bitácora general del sistema
class BitacoraListView(generics.ListAPIView):
    queryset = Bitacora.objects.all().order_by('-fecha')
    serializer_class = BitacoraSerializer
    permission_classes = [permissions.IsAuthenticated]


# 📂 Bitácora específica por empleado
class BitacoraEmpleadoAPIView(generics.ListAPIView):
    serializer_class = BitacoraEmpleadoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        empleado_id = self.kwargs.get("empleado_id")
        return BitacoraEmpleado.objects.filter(empleado_id=empleado_id).order_by('-fecha')